#!/usr/bin/env python3
"""backend/tables.py — 절 안의 표를 '엑셀형 그리드'로 읽고/쓰고/엑셀 왕복한다.

8장(연구개발비)처럼 표 중심 절을 위해, hwpx 표를 행×열 셀 그리드로 노출한다.
셀 편집은 곧바로 yaml/section_*.yaml 의 해당 cell_para 노드(text)에 반영되고,
[hwpx 빌드] 시 yaml2hwpx 오버레이로 **최종 hwpx 의 표 셀**로 나온다.

좌표계는 hwpx-yaml 파이프라인과 동일:
  표 path = sX/pY/tZ , 셀 문단 path = sX/pY/tZ/rR/cC/pW
한 셀에 문단이 여러 개면 그 셀 값은 문단 text 들을 줄바꿈으로 이어 보이고,
저장 시 줄 단위로 각 문단에 되돌려 쓴다(문단 수 유지).
"""
from __future__ import annotations

import copy
import itertools
import shutil
import sys
from pathlib import Path
from typing import Any

from . import config, pipeline, store

# 구조 편집(행/열)은 python-hwpx 표 API 로는 불가 → OWPML XML 을 직접 조작한다.
if str(config.SKILL_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(config.SKILL_SCRIPTS))
import hwpx_common as _H  # noqa: E402  (skill 공유 모듈)
from lxml import etree  # noqa: E402

from .tree import build_tree  # noqa: E402  (yaml → tree.json 재생성)

_NS = "{http://www.hancom.co.kr/hwpml/2011/paragraph}"


def _q(e) -> str:
    return etree.QName(e).localname


def _trs(el):
    return [c for c in el if _q(c) == "tr"]


def _tcs(tr):
    return [c for c in tr if _q(c) == "tc"]


def _addr(tc):
    a = tc.find(f"{_NS}cellAddr")
    return (int(a.get("colAddr")), int(a.get("rowAddr"))) if a is not None else (0, 0)


def _span(tc):
    s = tc.find(f"{_NS}cellSpan")
    return (int(s.get("colSpan", 1)), int(s.get("rowSpan", 1))) if s is not None else (1, 1)


_ID_SEED = itertools.count(970000001)


def _regen_ids(node) -> None:
    for c in node.iter():
        if c.get("id") is not None:
            c.set("id", str(next(_ID_SEED)))


def _blank_text(tc) -> None:
    for t in tc.iter(f"{_NS}t"):
        t.text = ""


def _sidx_of(table_path: str) -> int:
    return int(table_path.split("/")[0][1:])


# ── 그리드 조립 ──────────────────────────────────────────────────────────────
def _cells_of_table(index: dict[str, dict], tpath: str) -> list[dict]:
    """table path 아래 cell_para 들을 (row,col) 셀로 묶어 그리드 셀 목록 반환."""
    prefix = tpath + "/"
    by_cell: dict[tuple[int, int], list[dict]] = {}
    for p, n in index.items():
        if not p.startswith(prefix) or n.get("kind") != "cell_para":
            continue
        key = (int(n.get("row", 0)), int(n.get("col", 0)))
        by_cell.setdefault(key, []).append(n)

    cells: list[dict] = []
    for (r, c), paras in sorted(by_cell.items()):
        paras.sort(key=lambda n: n.get("path", ""))          # p0, p1, …
        span = paras[0].get("span") or [1, 1]
        text = "\n".join((pp.get("text") or "") for pp in paras)
        cells.append({
            "row": r,
            "col": c,
            "rowspan": int(span[0]) if span else 1,
            "colspan": int(span[1]) if len(span) > 1 else 1,
            "paths": [pp.get("path") for pp in paras],
            "text": text,
        })
    return cells


def tables_for(pid: str, nid: str) -> dict:
    """절 nid 의 모든 표를 그리드로. {nid, has_tables, tables:[{path,rows,cols,cells}]}"""
    node = store.node_by_id(pid, nid) or {}
    table_paths = list(node.get("table_paths", []) or [])
    index = pipeline._all_nodes_by_path(pid)

    tables: list[dict] = []
    for i, tp in enumerate(table_paths):
        tnode = index.get(tp, {})
        tables.append({
            "index": i,
            "path": tp,
            "rows": int(tnode.get("rows", 0) or 0),
            "cols": int(tnode.get("cols", 0) or 0),
            "cells": _cells_of_table(index, tp),
        })
    return {
        "nid": nid,
        "label": node.get("label", ""),
        "title": node.get("title", ""),
        "has_tables": bool(tables),
        "table_count": len(tables),
        "tables": tables,
    }


# ── 작성 힌트: 이 절 기존 표의 열 구조 → 프롬프트용 ───────────────────────────
import re as _re  # noqa: E402

# doc_fill 이 제반사항(overview)으로 채우는 정형 표는 AI 가 만들 대상이 아니므로 힌트에서 제외.
_DOCFILL_TABLE_SIGS = (
    ("연구개발과제명", "주관연구개발기관"),   # 표지
    ("국문핵심어",),                          # 요약문
    ("담당기술내용", "참여연구원"),           # 3-3 편성도
    ("기관부담연구개발비", "비율"),           # 8-1 지원·부담계획
)


def _norm_txt(s: str) -> str:
    return _re.sub(r"[^0-9A-Za-z가-힣]", "", (s or ""))


def _is_docfill_table(norm_blob: str) -> bool:
    return any(all(_norm_txt(t) in norm_blob for t in sig)
               for sig in _DOCFILL_TABLE_SIGS)


def _header_columns(cells: list[dict]) -> list[str]:
    """표의 머리행(row0, colspan 아래 row1 소제목 포함)에서 열 라벨 목록을 만든다."""
    row1 = {c["col"]: c for c in cells if c["row"] == 1}
    cols: list[str] = []
    for cell in sorted((c for c in cells if c["row"] == 0), key=lambda x: x["col"]):
        label = (cell.get("text") or "").replace("\n", " ").strip()
        span = int(cell.get("colspan", 1) or 1)
        subs: list[str] = []
        if span > 1:
            for cc in range(cell["col"], cell["col"] + span):
                s = row1.get(cc)
                st = (s.get("text") or "").replace("\n", " ").strip() if s else ""
                if st:
                    subs.append(st)
        if subs:
            label = f"{label}({'/'.join(subs)})" if label else "/".join(subs)
        cols.append(label or "-")
    return cols


def table_schema_hint(pid: str, nid: str, *, max_tables: int = 8,
                      max_chars: int = 2000) -> str:
    """이 절 문서에 이미 있는 '내용 표'의 열 구조를 프롬프트 힌트 문자열로.

    작성요령(제출 시 삭제)·doc_fill 정형 표(표지·요약문·편성도·8-1)는 제외한다.
    AI 가 마크다운 표를 그 열 구조에 맞춰 쓰면 빌드 시 기존 표에 정확히 흡수된다."""
    try:
        data = tables_for(pid, nid)
    except Exception:  # noqa: BLE001
        return ""
    if not data.get("has_tables"):
        return ""
    lines: list[str] = []
    for t in data["tables"]:
        cells = t.get("cells") or []
        blob = _norm_txt(" ".join(c.get("text") or "" for c in cells))
        if "작성요령" in blob or _is_docfill_table(blob):
            continue
        cols = [c for c in _header_columns(cells) if c and c != "-"]
        if len(cols) < 2:
            continue
        # 머리글이 문단 길이(설명·참고자료 표)면 채울 표가 아니므로 제외.
        if any(len(c) > 40 for c in cols):
            continue
        lines.append(
            f"- 표{t['index'] + 1}({t.get('rows', 0)}행×{t.get('cols', 0)}열): "
            "| " + " | ".join(cols) + " |")
        if len(lines) >= max_tables:
            break
    if not lines:
        return ""
    body = "\n".join(lines)
    if len(body) > max_chars:
        body = body[:max_chars].rstrip() + " …"
    return (
        "[기존 표 양식(이 절 문서에 이미 있는 표)]\n"
        "표 형태의 내용을 쓸 때는 **아래 표의 열 머리글·순서를 그대로 따라** 마크다운 표"
        "(| 열1 | 열2 | … |)로 작성하라. 열을 추가·삭제·재배치하지 말 것. 빌드 시 이 마크다운 "
        "표가 아래 기존 표에 자동으로 채워진다(중복 표를 새로 만들지 않는다).\n" + body)


# ── 한꺼번에 표 채우기: 작성된 마크다운 표 → 기존 표에 흡수(소스에 bake) ────────
def fill_all_tables(pid: str) -> dict:
    """모든 절의 작성 결과(마크다운 표)를 문서의 기존 표에 흡수(재구성)해 **웹 그리드·소스에
    영구 반영**한다.

    빌드와 동일한 흡수 로직(hwpx_common.fill_template_tables_from_markdown)을 표 전용 모드로
    실행한 hwpx 를 새 source 로 굽고(bake), yaml·tree 를 재생성한다. 서식(글꼴·출처취합 등)
    부작용 없이 표만 채운다. 실패 시 source·yaml 을 원상복구한다.

    반환 {ok, tables_before, tables_after, changed}."""
    base = _base_path(pid)
    ydir = store.yaml_dir(pid)
    outd = store.output_dir(pid)
    outd.mkdir(parents=True, exist_ok=True)
    tmp = outd / "_filltab.hwpx"
    bak_base = outd / "_bak_source.hwpx"
    bak_yaml = outd / "_bak_yaml"
    bak_tree = outd / "_bak_tree.json"
    tree_path = store._tree_json(pid)

    def _count_tables() -> int:
        idx = pipeline._all_nodes_by_path(pid)
        return sum(1 for n in idx.values() if n.get("kind") == "table")

    before = _count_tables()
    # 롤백 백업
    shutil.copy(str(base), str(bak_base))
    if bak_yaml.exists():
        shutil.rmtree(str(bak_yaml))
    shutil.copytree(str(ydir), str(bak_yaml))
    if tree_path.exists():
        shutil.copy(str(tree_path), str(bak_tree))
    try:
        # 1) 표 흡수만 켠 복원 → 마크다운 표가 기존 표에 채워진 hwpx
        pipeline.restore(str(base), str(ydir), str(tmp), only_tables=True)
        # 2) 새 source 로 굽고 yaml 재추출 + tree 재생성(제목 기반 id 안정 → 내비 유지)
        shutil.copy(str(tmp), str(base))
        pipeline.extract(str(base), str(ydir))
        store._write_json(tree_path, build_tree(str(ydir)))
        after = _count_tables()
        return {"ok": True, "tables_before": before, "tables_after": after,
                "changed": before != after}
    except Exception as exc:  # noqa: BLE001 - 실패하면 원상복구
        shutil.copy(str(bak_base), str(base))
        if ydir.exists():
            shutil.rmtree(str(ydir))
        shutil.copytree(str(bak_yaml), str(ydir))
        if bak_tree.exists():
            shutil.copy(str(bak_tree), str(tree_path))
        raise RuntimeError(f"표 채우기 실패(원상복구됨): ({type(exc).__name__}) {exc}") from exc


# ── 저장(그리드 → yaml) ──────────────────────────────────────────────────────
def _spread(text: str, paths: list[str]) -> list[dict]:
    """셀 값(줄바꿈 포함)을 셀의 문단 path 들에 줄 단위로 분배."""
    if not paths:
        return []
    lines = (text or "").split("\n")
    out: list[dict] = []
    n = len(paths)
    for i, p in enumerate(paths):
        if i < n - 1:
            val = lines[i] if i < len(lines) else ""
        else:  # 마지막 문단: 남은 줄을 합쳐 담아 내용 손실 방지
            val = "\n".join(lines[i:]) if i < len(lines) else ""
        out.append({"path": p, "text": val})   # marker 생략 → 기존 마커 보존
    return out


def save_cells(pid: str, nid: str, cells: list[dict]) -> dict:
    """cells = [{paths:[...], text}] (변경된 셀만). yaml 에 반영하고 통계 반환."""
    result: list[dict] = []
    for cell in cells or []:
        paths = cell.get("paths") or ([cell["path"]] if cell.get("path") else [])
        result.extend(_spread(cell.get("text", ""), paths))
    stats = pipeline.merge_result_into_yaml(pid, result)
    stats["cells"] = len(cells or [])
    return stats


# ── 엑셀 내보내기/가져오기 ────────────────────────────────────────────────────
def to_xlsx(pid: str, nid: str, out_path: Path) -> Path:
    """절의 표들을 .xlsx 로. 표마다 시트 1개, 셀 위치 그대로. 병합 셀 반영."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = tables_for(pid, nid)
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="center")

    for t in data["tables"]:
        title = f"{data['label'] or nid}_표{t['index'] + 1}"[:31]
        ws = wb.create_sheet(title=title)
        for cell in t["cells"]:
            r, c = cell["row"] + 1, cell["col"] + 1
            wc = ws.cell(row=r, column=c, value=cell["text"])
            wc.alignment = wrap
            if cell["row"] == 0:
                wc.font = bold
                wc.fill = header_fill
            rs, cs = cell.get("rowspan", 1), cell.get("colspan", 1)
            if rs > 1 or cs > 1:
                ws.merge_cells(start_row=r, start_column=c,
                               end_row=r + rs - 1, end_column=c + cs - 1)
        # 열 폭 살짝 넓게
        for col in range(1, (t["cols"] or 1) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
    if not data["tables"]:
        wb.create_sheet(title="빈표")
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def from_xlsx(pid: str, nid: str, xlsx_path: Path) -> dict:
    """업로드된 .xlsx 를 위치 기준으로 그리드에 되읽어 저장한다.
    시트는 순서(index)로, 셀은 (row+1,col+1) 위치로 매칭한다."""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), data_only=True)
    sheets = wb.worksheets
    data = tables_for(pid, nid)

    edits: list[dict] = []
    for t in data["tables"]:
        if t["index"] >= len(sheets):
            break
        ws = sheets[t["index"]]
        for cell in t["cells"]:
            v = ws.cell(row=cell["row"] + 1, column=cell["col"] + 1).value
            new = "" if v is None else str(v)
            if new != cell["text"]:
                edits.append({"paths": cell["paths"], "text": new})
    stats = save_cells(pid, nid, edits)
    stats["imported_sheets"] = min(len(sheets), len(data["tables"]))
    return stats


# ── 구조 편집(행/열 추가·삭제) ────────────────────────────────────────────────
# 편집은 프로젝트의 base(source.hwpx)를 진화시킨다:
#   1) 현재 yaml 텍스트를 base 에 얹어(restore) 최신 상태 hwpx 를 만든다
#   2) 그 hwpx 의 표 XML 을 조작(행/열)한다
#   3) base 를 교체하고 yaml 을 재추출해 좌표계를 다시 맞춘다
def _base_path(pid: str) -> Path:
    return store.project_dir(pid) / "source.hwpx"


def _apply_structural(pid: str, table_path: str, mutate, pure: bool = False) -> dict:
    base = _base_path(pid)
    ydir = store.yaml_dir(pid)
    outd = store.output_dir(pid)
    outd.mkdir(parents=True, exist_ok=True)
    tmp = outd / "_struct.hwpx"
    # 롤백 백업(구조편집 실패 시 프로젝트 원상복구)
    bak_base = outd / "_bak_source.hwpx"
    bak_yaml = outd / "_bak_yaml"
    shutil.copy(str(base), str(bak_base))
    if bak_yaml.exists():
        shutil.rmtree(str(bak_yaml))
    shutil.copytree(str(ydir), str(bak_yaml))
    try:
        # pure=True: 후처리 없는 순수 복원(문단 좌표 보존 → resolve_table 정확)
        pipeline.restore(str(base), str(ydir), str(tmp), pure=pure)   # 1) 현재 텍스트 반영
        doc = _H.open_hwpx(str(tmp))                        # 2) 표 XML 조작
        tbl = _H.resolve_table(doc, table_path)
        info = mutate(tbl)
        doc.sections[_sidx_of(table_path)].mark_dirty()
        _H.save_hwpx(doc, str(tmp))
        shutil.copy(str(tmp), str(base))                   # 3) base 교체 + yaml 재추출
        pipeline.extract(str(base), str(ydir))             # (그리드 검증도 겸함)
        return info or {}
    except Exception as exc:  # noqa: BLE001 - 실패하면 원상복구
        shutil.copy(str(bak_base), str(base))
        if ydir.exists():
            shutil.rmtree(str(ydir))
        shutil.copytree(str(bak_yaml), str(ydir))
        raise RuntimeError(
            f"구조 편집 실패(원상복구됨): 병합셀이 많은 표는 이 위치에서 행/열 변경이 "
            f"어려울 수 있습니다. ({type(exc).__name__})"
        ) from exc


# ── 8-1형(단계/연차/기관) 표: 단계 수에 맞춰 단계 블록 재구성 ─────────────────
def _cell_text(tc) -> str:
    t = tc.find(f".//{_NS}t")
    return (t.text or "") if t is not None else ""


def _set_cell_text(tc, s: str) -> None:
    t = tc.find(f".//{_NS}t")
    if t is not None:
        t.text = s


def _c0_text(tr) -> str:
    for tc in _tcs(tr):
        if _addr(tc)[0] == 0:
            return _cell_text(tc).strip()
    return ""


def _is_stage_anchor(tr) -> bool:
    t = _c0_text(tr)
    return t.isdigit() or t in ("n", "N")


def _set_stage_label(tr, label: str) -> None:
    for tc in _tcs(tr):
        if _addr(tc)[0] == 0:
            _set_cell_text(tc, label)
            return


def _tc_at(tr_or_cells, col: int):
    cells = _tcs(tr_or_cells) if hasattr(tr_or_cells, "iter") else tr_or_cells
    for tc in cells:
        if _addr(tc)[0] == col:
            return tc
    return None


def _set_rowspan(tc, n: int) -> None:
    s = tc.find(f"{_NS}cellSpan")
    if s is not None:
        s.set("rowSpan", str(n))


def rebuild_budget_stages(pid: str, table_path: str, stage_year_counts) -> dict:
    """8-1형 표를 '단계 × 연차 × 기관' 완전 구조로 재구성한다.

    - **연차1 서브그룹(기관별 데이터행+비율행)** 을 원자 단위로 삼아, 모든 연차를 그 구조로 만든다
      (=연차2 이후에도 비율행 생성). 각 단계는 (Y 연차 유닛 + 소계) 로 구성, 마지막에 총계.
    - `stage_year_counts` = 단계별 연차 수 리스트(예 [2,2] = 2단계 각 2연차).
    - 데이터 셀은 비우고(이후 doc_fill 채움), 단계/연차 라벨·병합(rowSpan)·rowAddr·rowCnt 를 재설정.
    구조편집이므로 pure 복원(후처리 OFF)으로 좌표를 맞춘 뒤 수행하며, 실패 시 롤백된다."""
    counts = [max(1, int(c)) for c in (stage_year_counts or [1])] or [1]

    def mutate(tbl):
        el = tbl.element
        trs = _trs(el)
        # 앵커 인식
        s1_idx = next((i for i, tr in enumerate(trs) if _is_stage_anchor(tr)), None)
        total_idx = next((i for i, tr in enumerate(trs) if _c0_text(tr) == "총계"), None)
        if s1_idx is None or total_idx is None:
            return {"ok": False, "reason": "구조 인식 실패"}
        # 연차1 유닛 높이 = 단계1 첫 행의 c1(연차) rowSpan
        c1_cell = _tc_at(trs[s1_idx], 1)
        yh = _span(c1_cell)[1] if c1_cell is not None else 8   # 유닛 행 수(기관수×2)
        # 소계 tr(단계1 범위 안 c1='소계')
        soke_idx = next((i for i in range(s1_idx, total_idx)
                         if _c1_text(trs[i]) == "소계"), None)
        if soke_idx is None:
            return {"ok": False, "reason": "소계행 없음"}
        header = trs[:s1_idx]
        year_unit = [copy.deepcopy(t) for t in trs[s1_idx:s1_idx + yh]]  # 연차 유닛 원본
        soke_tpl = copy.deepcopy(trs[soke_idx])
        total_tpl = copy.deepcopy(trs[total_idx])

        def _blank_unit(unit):
            for tr in unit:
                for tc in _tcs(tr):
                    if _addr(tc)[0] in (2, 3, 4, 5, 6, 7, 8, 9, 10):
                        _blank_text(tc)

        def _blank_sums(tr):
            for tc in _tcs(tr):
                if _addr(tc)[0] in (4, 5, 6, 7, 8, 9, 10):
                    _blank_text(tc)

        # 새 본문 조립
        new_body = []
        for k, Y in enumerate(counts, start=1):
            stage_rows = Y * yh + 1  # 연차 유닛들 + 소계 1행
            for j in range(Y):
                unit = [copy.deepcopy(t) for t in year_unit]
                _regen_ids_list(unit)
                _blank_unit(unit)
                first = unit[0]
                # 연차 라벨
                c1c = _tc_at(first, 1)
                if c1c is not None:
                    _set_cell_text(c1c, str(j + 1))
                c0c = _tc_at(first, 0)
                if j == 0:
                    # 단계 라벨 + 단계 전체 rowSpan
                    if c0c is not None:
                        _set_cell_text(c0c, str(k))
                        _set_rowspan(c0c, stage_rows)
                else:
                    # 후속 연차: 단계 셀 제거(위 단계 셀 rowSpan 이 덮음)
                    if c0c is not None:
                        first.remove(c0c)
                new_body.extend(unit)
            soke = copy.deepcopy(soke_tpl)
            _regen_ids(soke)
            _blank_sums(soke)
            # 소계행의 c0(단계) 셀은 없어야 함(단계 rowSpan 이 덮음)
            sc0 = _tc_at(soke, 0)
            if sc0 is not None:
                soke.remove(sc0)
            new_body.append(soke)
        total = copy.deepcopy(total_tpl)
        _regen_ids(total)
        _blank_sums(total)
        new_body.append(total)

        # 본문 교체: 헤더 뒤 모든 tr 제거 후 새 본문 추가
        for tr in trs[s1_idx:]:
            el.remove(tr)
        for tr in new_body:
            el.append(tr)
        # rowAddr 재부여(행 순번)
        all_trs = _trs(el)
        for i, tr in enumerate(all_trs):
            for tc in _tcs(tr):
                a = tc.find(f"{_NS}cellAddr")
                if a is not None:
                    a.set("rowAddr", str(i))
        el.set("rowCnt", str(len(all_trs)))
        return {"ok": True, "stages": len(counts), "rows": len(all_trs)}

    return _apply_structural(pid, table_path, mutate, pure=True)


# ── 8-3 비목별 세부표: 표 개수를 참여기관 수에 맞춰 복제/제거(구조편집) ────────
def _apply_structural_doc(pid: str, mutate, pure: bool = False) -> dict:
    """`_apply_structural` 의 doc 단위 버전 — mutate(doc) 가 문서 전체(섹션/문단)를 조작.

    표 하나가 아니라 '표를 품은 문단' 자체를 복제/제거할 때 쓴다. 실패 시 롤백."""
    base = _base_path(pid)
    ydir = store.yaml_dir(pid)
    outd = store.output_dir(pid)
    outd.mkdir(parents=True, exist_ok=True)
    tmp = outd / "_struct.hwpx"
    bak_base = outd / "_bak_source.hwpx"
    bak_yaml = outd / "_bak_yaml"
    shutil.copy(str(base), str(bak_base))
    if bak_yaml.exists():
        shutil.rmtree(str(bak_yaml))
    shutil.copytree(str(ydir), str(bak_yaml))
    try:
        pipeline.restore(str(base), str(ydir), str(tmp), pure=pure)
        doc = _H.open_hwpx(str(tmp))
        info = mutate(doc) or {}
        for si in (info.get("sections") or range(len(doc.sections))):
            doc.sections[si].mark_dirty()
        _H.save_hwpx(doc, str(tmp))
        shutil.copy(str(tmp), str(base))
        pipeline.extract(str(base), str(ydir))
        return info
    except Exception as exc:  # noqa: BLE001 - 실패하면 원상복구
        shutil.copy(str(bak_base), str(base))
        if ydir.exists():
            shutil.rmtree(str(ydir))
        shutil.copytree(str(bak_yaml), str(ydir))
        raise RuntimeError(f"구조 편집 실패(원상복구됨): ({type(exc).__name__}) {exc}") from exc


def rebuild_budget_detail(pid: str, target_n: int, detail_paths: list[str]) -> dict:
    """8-3 비목별 세부표를 참여기관 수(target_n)에 맞춰 복제/제거한다.

    detail_paths = 현재 세부표 path 목록(문서순, doc_fill 시그니처로 탐색). 세부표를 품은
    섹션 직속 문단을 lxml deepcopy 로 복제(마지막 세부표를 템플릿으로 뒤에 추가)하거나,
    남는 것은 뒤에서 제거한다. pure 복원으로 좌표를 맞춘 뒤 수행하며 실패 시 롤백."""
    target_n = max(1, int(target_n))

    def mutate(doc):
        hosts = []
        touched: set[int] = set()
        for tp in detail_paths:
            try:
                tbl = _H.resolve_table(doc, tp)
            except Exception:  # noqa: BLE001 - 못 찾으면 건너뜀
                continue
            hp = _H._host_para_of(tbl.element)
            if hp is not None:
                hosts.append(hp)
                touched.add(_sidx_of(tp))
        if not hosts:
            return {"ok": False, "reason": "세부표 없음", "before": 0, "after": 0}
        cur = len(hosts)
        if target_n > cur:
            anchor = hosts[-1]
            for _ in range(target_n - cur):
                clone = copy.deepcopy(hosts[-1])   # 마지막 세부표(템플릿) 복제
                _regen_ids(clone)
                anchor.addnext(clone)
                anchor = clone
        elif target_n < cur:
            for hp in hosts[target_n:]:
                par = hp.getparent()
                if par is not None:
                    par.remove(hp)
        return {"ok": True, "before": cur, "after": target_n, "sections": sorted(touched)}

    return _apply_structural_doc(pid, mutate, pure=True)


# ── 7-1 공동연구개발기관책임자 프로필 블록: 공동기관 수만큼 복제/제거 ──────────
def rebuild_researcher_blocks(pid: str, target_n: int) -> dict:
    """7-1 '공동연구개발기관책임자' 프로필 블록을 공동기관 수(target_n)에 맞춰 복제/제거한다.

    한 블록 = 제목 문단 + 인적사항/학력/경력/국가R&D실적/논문·저서/지식재산권/수상/국외지원
    표 묶음(소제목·표·빈 문단 포함). 블록 경계는 '공동연구개발기관책임자' 제목 문단부터
    '참여연구자'(총괄) 문단 직전까지(경계 판정은 표를 품지 않은 순수 제목 문단 텍스트로만).
    섹션 직속 문단 구간을 lxml deepcopy 로 통째 복제(마지막 블록=템플릿, 앞으로 붙임)하거나
    초과분을 제거한다. pure 복원으로 좌표를 맞춘 뒤 수행하며 실패 시 롤백. 값은 비운 채(템플릿
    그대로) 두어 사용자가 그리드에서 직접 채운다. 성공 시 새 표가 ③ 입력 그리드에 뜨도록
    tree.json(절별 table_paths)을 재생성한다."""
    target_n = max(1, int(target_n))     # 최소 1(공동 템플릿 블록 보존 — 삭제 시 재추가 불가)

    def mutate(doc):
        for si, sec in enumerate(doc.sections):
            kids = [c for c in list(sec.element) if c.tag.endswith("}p")]
            # 경계·헤딩 판정: 표를 품은 문단은 셀 텍스트 오탐 방지를 위해 "" 처리.
            texts = ["" if _H._has_table_el(p) else _norm_txt(_H._p_text_el(p))
                     for p in kids]
            start = next((i for i, t in enumerate(texts)
                          if "공동연구개발기관책임자" in t), None)
            if start is None:
                continue
            bound = next((i for i in range(start, len(texts))
                          if "참여연구자" in texts[i]), None)
            if bound is None:
                return {"ok": False, "reason": "참여연구자 총괄 경계 없음",
                        "before": 0, "after": 0}
            heads = [i for i in range(start, bound)
                     if "공동연구개발기관책임자" in texts[i]]
            cur = len(heads)
            boundary_el = kids[bound]
            if target_n > cur:
                tpl = kids[heads[-1]:bound]            # 마지막 블록 = 복제 템플릿
                for _ in range(target_n - cur):
                    for e in tpl:
                        clone = copy.deepcopy(e)
                        _regen_ids(clone)
                        boundary_el.addprevious(clone)   # 참여연구자 총괄 직전에 순서대로 삽입
            elif target_n < cur:
                for e in kids[heads[target_n]:bound]:  # 초과 블록(문단 구간) 제거
                    par = e.getparent()
                    if par is not None:
                        par.remove(e)
            return {"ok": True, "before": cur, "after": target_n, "sections": [si]}
        return {"ok": False, "reason": "공동책임자 블록 없음", "before": 0, "after": 0}

    info = _apply_structural_doc(pid, mutate, pure=True)
    if info.get("ok") and info.get("before") != info.get("after"):
        # 표 개수가 바뀌었으니 절별 table_paths(tree.json)를 재생성 → 그리드 즉시 반영.
        store._write_json(store._tree_json(pid), build_tree(str(store.yaml_dir(pid))))
    return info


def _c1_text(tr) -> str:
    for tc in _tcs(tr):
        if _addr(tc)[0] == 1:
            return _cell_text(tc).strip()
    return ""


def _regen_ids_list(trs) -> None:
    for tr in trs:
        _regen_ids(tr)


def _row_owner_tr(el, row: int):
    """논리 행 row 를 '시작'으로 갖는 tr(행 삽입/삭제의 기준)."""
    for tr in _trs(el):
        cells = _tcs(tr)
        if cells and min(_addr(c)[1] for c in cells) == row:
            return tr
    trs = _trs(el)
    return trs[row] if 0 <= row < len(trs) else (trs[-1] if trs else None)


def add_row(pid: str, nid: str, table_path: str, after_row: int) -> dict:
    """논리 행 after_row 아래에 행 하나 삽입(세로병합 인지).
    - 삽입선을 가로지르는 세로병합 셀 → rowSpan +1
    - after_row 에서 끝나는 셀만 새 행에 복제(서식 유지, 텍스트 비움)
    - after_row 아래 행들은 rowAddr +1"""
    def mutate(tbl):
        el = tbl.element
        at = after_row
        ins = at + 1
        trs = _trs(el)

        def _owner(row):
            for tr in trs:
                cs = _tcs(tr)
                if cs and min(_addr(c)[1] for c in cs) == row:
                    return tr
            return None

        anchor_tr = _owner(ins)          # 삽입 위치(옛 at+1) 앞에 새 tr
        owner_at = _owner(at)
        clones = []
        for tr in trs:
            for tc in _tcs(tr):
                col, row = _addr(tc)
                csp, rsp = _span(tc)
                bot = row + rsp - 1
                if row <= at and bot >= ins:            # 삽입선 가로지름 → 확장
                    sp = tc.find(f"{_NS}cellSpan")
                    sp.set("rowSpan", str(rsp + 1))
                elif row >= ins:                        # 아래로 밀기
                    a = tc.find(f"{_NS}cellAddr")
                    a.set("rowAddr", str(row + 1))
                if bot == at:                           # at 에서 끝나는 셀 → 새 행 복제
                    clones.append(tc)

        new_tr = etree.Element(f"{_NS}tr")
        for tc in sorted(clones, key=lambda t: _addr(t)[0]):
            nc = copy.deepcopy(tc)
            _regen_ids(nc)
            nc.find(f"{_NS}cellAddr").set("rowAddr", str(ins))
            sp = nc.find(f"{_NS}cellSpan")
            if sp is not None:
                sp.set("rowSpan", "1")
            _blank_text(nc)
            new_tr.append(nc)
        if anchor_tr is not None:
            anchor_tr.addprevious(new_tr)
        elif owner_at is not None:
            owner_at.addnext(new_tr)
        elif trs:
            trs[-1].addnext(new_tr)
        else:
            el.append(new_tr)
        el.set("rowCnt", str(int(el.get("rowCnt", "0")) + 1))
        return {"rows": int(el.get("rowCnt"))}
    return _apply_structural(pid, table_path, mutate)


def delete_row(pid: str, nid: str, table_path: str, row: int) -> dict:
    """논리 행 row 삭제(세로병합 인지). 가로지르는 셀은 rowSpan -1, 그 행 전용 셀은 제거."""
    def mutate(tbl):
        el = tbl.element
        for tr in list(_trs(el)):
            for tc in list(_tcs(tr)):
                col, r = _addr(tc)
                csp, rsp = _span(tc)
                bot = r + rsp - 1
                if r == row and rsp == 1:               # 이 행 전용 셀 → 제거
                    tr.remove(tc)
                elif r < row <= bot:                    # 가로지름 → 축소
                    tc.find(f"{_NS}cellSpan").set("rowSpan", str(rsp - 1))
                elif r == row and rsp > 1:              # 이 행에서 시작+아래로 → 축소+한칸 아래
                    tc.find(f"{_NS}cellSpan").set("rowSpan", str(rsp - 1))
                    tc.find(f"{_NS}cellAddr").set("rowAddr", str(row + 1))
                elif r > row:                           # 위로 당기기
                    tc.find(f"{_NS}cellAddr").set("rowAddr", str(r - 1))
            if not _tcs(tr):
                el.remove(tr)
        el.set("rowCnt", str(max(0, int(el.get("rowCnt", "1")) - 1)))
        return {"rows": int(el.get("rowCnt"))}
    return _apply_structural(pid, table_path, mutate)


def _tc_at(tr, col: int):
    for tc in _tcs(tr):
        c = _addr(tc)[0]
        cs = _span(tc)[0]
        if c <= col < c + cs:
            return tc
    return None


def add_col(pid: str, nid: str, table_path: str, after_col: int) -> dict:
    """열 하나 삽입(가로병합 인지). 삽입선을 가로지르는 가로병합 셀은 colSpan +1,
    after_col 에서 끝나는 셀만 새 열에 복제."""
    def mutate(tbl):
        el = tbl.element
        at = after_col
        ins = at + 1
        for tr in _trs(el):
            clones = []
            for tc in _tcs(tr):
                col = _addr(tc)[0]
                csp = _span(tc)[0]
                right = col + csp - 1
                if col <= at and right >= ins:          # 가로지름 → 확장
                    tc.find(f"{_NS}cellSpan").set("colSpan", str(csp + 1))
                elif col >= ins:                        # 오른쪽으로 밀기
                    tc.find(f"{_NS}cellAddr").set("colAddr", str(col + 1))
                if right == at:                         # at 에서 끝나는 셀 → 복제
                    clones.append(tc)
            for tc in clones:
                nc = copy.deepcopy(tc)
                _regen_ids(nc)
                nc.find(f"{_NS}cellAddr").set("colAddr", str(ins))
                sp = nc.find(f"{_NS}cellSpan")
                if sp is not None:
                    sp.set("colSpan", "1")
                _blank_text(nc)
                tc.addnext(nc)
        el.set("colCnt", str(int(el.get("colCnt", "0")) + 1))
        return {"cols": int(el.get("colCnt"))}
    return _apply_structural(pid, table_path, mutate)


def delete_col(pid: str, nid: str, table_path: str, col: int) -> dict:
    """열 col 삭제(가로병합 인지)."""
    def mutate(tbl):
        el = tbl.element
        for tr in _trs(el):
            for tc in list(_tcs(tr)):
                c = _addr(tc)[0]
                csp = _span(tc)[0]
                right = c + csp - 1
                if c == col and csp == 1:
                    tr.remove(tc)
                elif c < col <= right:
                    tc.find(f"{_NS}cellSpan").set("colSpan", str(csp - 1))
                elif c == col and csp > 1:
                    tc.find(f"{_NS}cellSpan").set("colSpan", str(csp - 1))
                    tc.find(f"{_NS}cellAddr").set("colAddr", str(col + 1))
                elif c > col:
                    tc.find(f"{_NS}cellAddr").set("colAddr", str(c - 1))
        el.set("colCnt", str(max(0, int(el.get("colCnt", "1")) - 1)))
        return {"cols": int(el.get("colCnt"))}
    return _apply_structural(pid, table_path, mutate)


# ── 수식(합계 등) 사이드카 ────────────────────────────────────────────────────
# 셀 수식(=SUM(...), =A1+B1 …)은 프런트가 계산해 '값'을 hwpx 에 쓰고,
# 원본 수식은 여기(node/formulas.json)에 보관해 다시 열 때 편집할 수 있게 한다.
import json as _json


def _formulas_path(pid: str, nid: str) -> Path:
    return store.node_dir(pid, nid) / "formulas.json"


def get_formulas(pid: str, nid: str) -> dict:
    p = _formulas_path(pid, nid)
    try:
        return _json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_formulas(pid: str, nid: str, formulas: dict) -> dict:
    p = _formulas_path(pid, nid)
    p.parent.mkdir(parents=True, exist_ok=True)
    # 빈 값은 제거해 깔끔하게 유지
    clean = {k: v for k, v in (formulas or {}).items() if str(v or "").strip()}
    p.write_text(_json.dumps(clean, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"count": len(clean)}
